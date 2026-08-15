"""Derive contact_status for every target from the sponsor CRM. Manual, offline.

This script exists so nobody hand-types `known_contact` into targets.csv. It
reads the CRM, prints what it concluded and why, and writes three columns into
data/targets.csv only after you confirm.

It is deliberately NOT part of the pipeline. Phase 3 and Phase 4 read
targets.csv and never touch the workbook, so a live run cannot depend on a
spreadsheet being open, correct, or even present. tests/test_crm_isolation.py
enforces that.

Read-only, three ways: the workbook is opened with read_only=True, nothing in
here calls save(), and the file's size and mtime are printed before and after so
you can see it was not touched.

  contact_status
    cold_prospect     no partnership row anywhere in the CRM
    existing_partner  a partnership row whose expiration is in the future
    lapsed_partner    expired, or flagged non-renewing

  has_known_contact       TRUE if any tab holds a non-empty contact name
  contact_needs_refresh   FALSE on derivation. Only a human sets it TRUE.

Which tabs count for what:
  Partnership ledgers are tabs that carry an expiration column (the owner tabs
  and Archive). Only those decide contact_status. Pipeline & Leads is the lead
  list targets.csv was seeded from, so a row there is a lead, not a partnership.
  Every tab, ledger or not, counts toward has_known_contact.

Usage:
    python scripts/derive_target_status.py            # print the table, write nothing
    python scripts/derive_target_status.py --write    # print, then confirm, then write
"""

import argparse
import csv
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from common.config import load_settings, resolve_path  # noqa: E402
from common.logging import get_logger  # noqa: E402

log = get_logger("scripts.derive_target_status")

TARGETS_CSV = PROJECT_ROOT / "data" / "targets.csv"

DERIVED_COLUMNS = [
    "contact_status",
    "has_known_contact",
    "contact_needs_refresh",
    "relationship_record_id",
    "relationship_tier",
    "relationship_status",
    "relationship_contact_name",
    "relationship_contact_email",
    "relationship_expiration",
    "relationship_decline_reason",
    "relationship_crm_source",
]

COLD, EXISTING, LAPSED = "cold_prospect", "existing_partner", "lapsed_partner"

# A status that says the relationship is over, whatever the expiration says.
NON_RENEWING = {
    "not renewing", "non-renewing", "expired", "declined", "churned", "lost",
    "did not renew", "inactive",
}

# Placeholders that mean "no contact", which a truthiness test would miss.
EMPTY_MARKERS = {"", "-", "—", "–", "n/a", "na", "none", "tbd", "?", "unknown"}
FORBIDDEN_DASHES = str.maketrans({"—": ";", "–": ";", "―": ";"})

# Trailing words that describe a firm's legal form or asset class rather than
# its identity. Stripped from the end only: "Sixth Street Partners" and "Sixth
# Street" are the same firm, but "TD Bank" and "TD Securities" are not, so
# nothing is stripped from the middle and Bank/Securities are never stripped.
GENERIC_SUFFIXES = {
    "partners", "partner", "group", "capital", "management", "asset", "assets",
    "investments", "investment", "advisors", "advisers", "associates",
    "holdings", "international", "company", "co", "llp", "llc", "lp", "inc",
    "ltd", "plc", "the",
    # Left dangling once a suffix is stripped: "Piper Sandler & Co." becomes
    # "piper sandler and" without this.
    "and",
}

ABBREVIATIONS = {
    "mgmt": "management", "mgt": "management", "intl": "international",
    "int'l": "international", "&": "and", "corp": "corporation",
}

_PARENTHETICAL = re.compile(r"\([^)]*\)")
_PUNCTUATION = re.compile(r"[^a-z0-9 ]+")


# ── Firm name matching ────────────────────────────────────────────────────────

def normalize_firm(name: str) -> str:
    """Reduce a firm name to a comparison key.

    Conservative on purpose. A false match marks a cold prospect as an existing
    partner and silently suppresses a draft, so the derivation table prints the
    CRM row behind every match for you to check by eye.
    """
    text = _PARENTHETICAL.sub(" ", str(name or "")).lower()
    text = text.replace("&", " and ")
    text = _PUNCTUATION.sub(" ", text)
    words = [ABBREVIATIONS.get(w, w) for w in text.split()]
    while words and words[-1] in GENERIC_SUFFIXES:
        words.pop()
    while words and words[0] in GENERIC_SUFFIXES:
        words.pop(0)
    return " ".join(words)


def is_empty(value: Any) -> bool:
    return str(value or "").strip().lower() in EMPTY_MARKERS


def safe_crm_text(value: Any) -> str:
    """Copy CRM text into targets.csv without forbidden email punctuation.

    This performs punctuation normalization only. It does not summarize,
    soften, or otherwise compose CRM notes.
    """
    return " ".join(str(value or "").translate(FORBIDDEN_DASHES).split()).replace(" ;", ";")


# ── Expiration parsing ────────────────────────────────────────────────────────

_DATE_FORMATS = ("%b %d, %Y", "%B %d, %Y", "%b %d %Y", "%m/%d/%Y", "%Y-%m-%d",
                 "%d %b %Y")


def parse_expiration(value: Any) -> date | None:
    """Parse a CRM expiration cell. A bare year means the end of that year."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and 1900 < int(value) < 2200:
        return date(int(value), 12, 31)

    text = str(value).strip()
    if not text or text.lower() in EMPTY_MARKERS:
        return None
    if re.fullmatch(r"(19|20)\d{2}", text):
        return date(int(text), 12, 31)
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    # A date we cannot read is reported as unreadable, never guessed at.
    return None


# ── CRM reading ───────────────────────────────────────────────────────────────

def _header_row(rows: list[tuple], limit: int = 8) -> tuple[int, dict[str, int]] | None:
    """Find the header row and map lowercased header text to column index."""
    for index, row in enumerate(rows[:limit]):
        cells = {str(c).strip().lower(): i for i, c in enumerate(row) if c is not None}
        if "firm" in cells or "partnership" in cells:
            return index, cells
    return None


def _column(headers: dict[str, int], *candidates: str) -> int | None:
    """First header whose text contains any candidate substring."""
    for candidate in candidates:
        for text, index in headers.items():
            if candidate in text:
                return index
    return None


def _contact_columns(headers: dict[str, int]) -> list[int]:
    """Columns holding a person's name. Emails, dates, and flags are excluded."""
    return [
        index for text, index in headers.items()
        if "contact" in text
        and not any(skip in text for skip in ("email", "date", "verified", "on hand"))
    ]


class CrmRow:
    """One firm's row in one tab, reduced to what the derivation needs."""

    def __init__(self, tab: str, row_number: int, firm: str, status: str,
                 expiration_raw: Any, contacts: list[str], is_ledger: bool,
                 *, record_id: str = "", tier: str = "", emails: list[str] | None = None,
                 decline_reason: str = ""):
        self.tab = tab
        self.row_number = row_number
        self.firm = firm
        self.status = (status or "").strip()
        self.expiration_raw = expiration_raw
        self.expiration = parse_expiration(expiration_raw)
        self.contacts = [c for c in contacts if not is_empty(c)]
        self.record_id = safe_crm_text(record_id)
        self.tier = safe_crm_text(tier)
        self.emails = [safe_crm_text(e) for e in (emails or []) if not is_empty(e)]
        self.decline_reason = safe_crm_text(decline_reason)
        self.is_ledger = is_ledger

    @property
    def non_renewing(self) -> bool:
        return self.status.strip().lower() in NON_RENEWING

    def where(self) -> str:
        return f"{self.tab}!r{self.row_number}"

    def expiration_text(self) -> str:
        if self.expiration:
            return self.expiration.isoformat()
        raw = str(self.expiration_raw or "").strip()
        return f"unreadable({raw})" if raw else "none"


def read_crm(path: Path) -> tuple[list[CrmRow], list[str]]:
    """Read every tab. READ-ONLY: the workbook is never saved or modified.

    Returns (rows, skipped_tab_notes).
    """
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        rows: list[CrmRow] = []
        skipped: list[str] = []

        for tab in workbook.sheetnames:
            sheet = workbook[tab]
            grid = [tuple(r) for r in sheet.iter_rows(values_only=True)]
            found = _header_row(grid)
            if not found:
                skipped.append(f"{tab}: no Firm column, not read")
                continue

            header_index, headers = found
            firm_col = _column(headers, "firm", "partnership")
            status_col = _column(headers, "status", "stage")
            expiration_col = _column(headers, "expiration", "last cycle")
            id_col = _column(headers, "id")
            tier_col = _column(headers, "tier")
            email_col = _column(headers, "email")
            decline_reason_col = _column(headers, "renewal notes", "notes")
            contact_cols = _contact_columns(headers)
            # A tab that tracks expirations is a partnership ledger. A tab that
            # does not is a lead list or a tracker, and cannot make a firm a
            # partner.
            is_ledger = expiration_col is not None

            for offset, raw in enumerate(grid[header_index + 1:], start=header_index + 2):
                if firm_col is None or firm_col >= len(raw):
                    continue
                firm = raw[firm_col]
                if is_empty(firm) or str(firm).strip().lower() == "firm":
                    continue
                rows.append(CrmRow(
                    tab=tab,
                    row_number=offset,
                    firm=str(firm).strip(),
                    status=str(raw[status_col]).strip() if status_col is not None
                    and status_col < len(raw) and raw[status_col] is not None else "",
                    expiration_raw=raw[expiration_col] if expiration_col is not None
                    and expiration_col < len(raw) else None,
                    contacts=[str(raw[c]) for c in contact_cols
                              if c < len(raw) and raw[c] is not None],
                    is_ledger=is_ledger,
                    record_id=str(raw[id_col]) if id_col is not None
                    and id_col < len(raw) and raw[id_col] is not None else "",
                    tier=str(raw[tier_col]) if tier_col is not None
                    and tier_col < len(raw) and raw[tier_col] is not None else "",
                    emails=[str(raw[email_col])] if email_col is not None
                    and email_col < len(raw) and raw[email_col] is not None else [],
                    decline_reason=str(raw[decline_reason_col])
                    if decline_reason_col is not None
                    and decline_reason_col < len(raw)
                    and raw[decline_reason_col] is not None else "",
                ))
        return rows, skipped
    finally:
        workbook.close()


# ── Derivation ────────────────────────────────────────────────────────────────

class Derivation:
    """What was concluded for one target, and the evidence behind it."""

    def __init__(self, firm: str):
        self.firm = firm
        self.contact_status = COLD
        self.has_known_contact = False
        self.contact_needs_refresh = False
        self.matches: list[CrmRow] = []
        self.deciding: CrmRow | None = None
        self.notes: list[str] = []

    @property
    def evidence(self) -> str:
        if self.deciding:
            return f"{self.deciding.where()} [{self.deciding.status or 'no status'}]"
        if self.matches:
            return f"{self.matches[0].where()} (non-ledger)"
        return "no CRM row"


def derive(target_firm: str, crm_rows: list[CrmRow], today: date) -> Derivation:
    """Decide one firm's status. Ambiguity resolves away from cold, never toward it."""
    result = Derivation(target_firm)
    key = normalize_firm(target_firm)
    if not key:
        result.notes.append("firm name normalizes to nothing; left cold")
        return result

    result.matches = [r for r in crm_rows if normalize_firm(r.firm) == key]
    result.has_known_contact = any(r.contacts for r in result.matches)

    ledger = [r for r in result.matches if r.is_ledger]
    if not ledger:
        if result.matches:
            tabs = ", ".join(sorted({r.tab for r in result.matches}))
            result.notes.append(f"row(s) only in non-ledger tab(s): {tabs}")
        return result

    live = [r for r in ledger
            if r.expiration and r.expiration > today and not r.non_renewing]
    if live:
        # A current contract outranks an old archived one.
        result.deciding = max(live, key=lambda r: r.expiration)
        result.contact_status = EXISTING
        return result

    result.contact_status = LAPSED
    flagged = [r for r in ledger if r.non_renewing]
    dated = [r for r in ledger if r.expiration]
    if flagged:
        result.deciding = flagged[0]
        result.notes.append(f"flagged '{flagged[0].status}'")
    elif dated:
        result.deciding = max(dated, key=lambda r: r.expiration)
        result.notes.append(f"expired {result.deciding.expiration.isoformat()}")
    else:
        result.deciding = ledger[0]
        result.notes.append(
            "REVIEW: partnership row with no readable expiration, treated as lapsed "
            "so no cold draft is generated"
        )
    return result


# ── Table and write ───────────────────────────────────────────────────────────

def read_targets(path: Path) -> tuple[list[dict[str, str]], list[str]]:
    with path.open(newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        return list(reader), list(reader.fieldnames or [])


def print_table(rows: list[dict[str, str]], derivations: dict[str, Derivation],
                today: date) -> None:
    headers = ["firm", "contact_status", "known", "refresh", "expiration", "CRM evidence", "note"]
    widths = [28, 16, 5, 7, 16, 26, 44]

    def line(cells: Iterable[str]) -> str:
        return "  ".join(str(c)[:w].ljust(w) for c, w in zip(cells, widths)).rstrip()

    print(f"\nDerivation table  (today = {today.isoformat()})")
    print(line(headers))
    print(line("-" * w for w in widths))

    counts = {COLD: 0, EXISTING: 0, LAPSED: 0}
    for row in rows:
        d = derivations[row["firm"]]
        counts[d.contact_status] += 1
        print(line([
            d.firm,
            d.contact_status,
            "TRUE" if d.has_known_contact else "FALSE",
            "TRUE" if d.contact_needs_refresh else "FALSE",
            d.deciding.expiration_text() if d.deciding else "-",
            d.evidence,
            "; ".join(d.notes),
        ]))

    print(f"\n{len(rows)} target(s): "
          f"{counts[COLD]} cold_prospect, {counts[EXISTING]} existing_partner, "
          f"{counts[LAPSED]} lapsed_partner.")


def write_targets(path: Path, rows: list[dict[str, str]], fields: list[str],
                  derivations: dict[str, Derivation]) -> None:
    out_fields = fields + [c for c in DERIVED_COLUMNS if c not in fields]
    for row in rows:
        d = derivations[row["firm"]]
        owner = str(row.get("owner") or "").strip().lower()
        if owner not in {"jamari", "fola"}:
            raise ValueError(
                f"{row['firm']} has invalid owner {row.get('owner')!r}; "
                "expected 'jamari' or 'fola'."
            )
        row["owner"] = owner
        row["contact_status"] = d.contact_status
        row["has_known_contact"] = "TRUE" if d.has_known_contact else "FALSE"
        row["contact_needs_refresh"] = "TRUE" if d.contact_needs_refresh else "FALSE"
        deciding = d.deciding
        row["relationship_record_id"] = deciding.record_id if deciding else ""
        row["relationship_tier"] = deciding.tier if deciding else ""
        row["relationship_status"] = deciding.status if deciding else ""
        row["relationship_contact_name"] = "; ".join(deciding.contacts) if deciding else ""
        row["relationship_contact_email"] = "; ".join(deciding.emails) if deciding else ""
        row["relationship_expiration"] = (
            deciding.expiration.isoformat() if deciding and deciding.expiration else ""
        )
        row["relationship_decline_reason"] = deciding.decline_reason if deciding else ""
        row["relationship_crm_source"] = deciding.where() if deciding else ""
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=out_fields)
        writer.writeheader()
        writer.writerows(rows)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("--write", action="store_true",
                        help="Write the three columns into targets.csv after confirmation.")
    parser.add_argument("--yes", action="store_true",
                        help="Skip the prompt. Only for a run whose table you have read.")
    parser.add_argument("--today", help="Override today's date (YYYY-MM-DD), for testing.")
    args = parser.parse_args(argv)

    today = date.fromisoformat(args.today) if args.today else date.today()

    settings = load_settings()
    crm_path = resolve_path(settings["crm"]["xlsx"]["path"])
    if not crm_path.exists():
        print(f"CRM workbook not found: {crm_path}", file=sys.stderr)
        return 2

    before = crm_path.stat()
    crm_rows, skipped = read_crm(crm_path)
    print(f"Read {len(crm_rows)} CRM row(s) from {crm_path.name} (read-only).")
    for note in skipped:
        print(f"  skipped {note}")

    ledger_tabs = sorted({r.tab for r in crm_rows if r.is_ledger})
    other_tabs = sorted({r.tab for r in crm_rows if not r.is_ledger})
    print(f"  partnership ledgers (decide status): {', '.join(ledger_tabs)}")
    print(f"  other tabs (contact names only):     {', '.join(other_tabs)}")

    rows, fields = read_targets(TARGETS_CSV)
    derivations = {row["firm"]: derive(row["firm"], crm_rows, today) for row in rows}

    # An existing manual TRUE is preserved. Derivation never sets it, and never
    # destroys a human's override either.
    for row in rows:
        if str(row.get("contact_needs_refresh", "")).strip().upper() == "TRUE":
            derivations[row["firm"]].contact_needs_refresh = True
            derivations[row["firm"]].notes.append("manual refresh flag preserved")

    print_table(rows, derivations, today)

    after = crm_path.stat()
    unchanged = (before.st_mtime, before.st_size) == (after.st_mtime, after.st_size)
    print(f"CRM workbook {'unchanged' if unchanged else 'CHANGED'}: "
          f"size {after.st_size}, mtime {datetime.fromtimestamp(after.st_mtime):%Y-%m-%d %H:%M:%S}")

    if not args.write:
        print("\nNothing written. Re-run with --write to update data/targets.csv.")
        return 0

    if not args.yes:
        if not sys.stdin.isatty():
            print("\nRefusing to write without a terminal to confirm at. "
                  "Re-run with --yes once you have read the table.", file=sys.stderr)
            return 3
        answer = input('\nType "write" to update data/targets.csv: ').strip()
        if answer != "write":
            print("Not confirmed. Nothing written.")
            return 1

    write_targets(TARGETS_CSV, rows, fields, derivations)
    print(f"Wrote {len(rows)} row(s) and {len(DERIVED_COLUMNS)} derived column(s) to "
          f"data/targets.csv.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
