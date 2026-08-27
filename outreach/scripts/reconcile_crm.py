"""Dry-run-only reconciliation of the 2026-27 CRM workbook and Firm Library.

This command has no apply/import mode and contains no database write path.  It
reads the workbook in ``read_only`` mode, compares it with either the reviewed
``data/targets.csv`` seed snapshot or a caller-supplied JSON database export,
and writes CSV/JSON/Markdown review reports.

Usage:
    python scripts/reconcile_crm.py
    python scripts/reconcile_crm.py --database-json /tmp/targets-export.json
    python scripts/reconcile_crm.py --output-dir review/reconciliation
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import date, datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Iterable, Mapping

PROJECT_ROOT = Path(__file__).resolve().parent.parent
REPO_ROOT = PROJECT_ROOT.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from dashboard.crm import canonical_pipeline_stage, canonical_relationship  # noqa: E402
from scripts.derive_target_status import normalize_firm  # noqa: E402


DEFAULT_WORKBOOK = REPO_ROOT / "Sponsor_CRM_2026-27_UPDATED 8_7_26.xlsx"
DEFAULT_DATABASE_CSV = PROJECT_ROOT / "data" / "targets.csv"
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "review" / "reconciliation"

AUTHORITY = {
    "Explicit correction": 100,
    "Jamari": 90,
    "Fola": 80,
    "EMEA (Daniel & Hajar)": 80,
    "Pipeline & Leads": 70,
    "Archive": 60,
    "Jamari (Updated)": 50,
    "Follow-Up Tracker": 40,
    "Current database": 0,
}

CORE_SHEETS = tuple(name for name in AUTHORITY if name not in {
    "Explicit correction", "Current database"
})

FIELD_HEADERS = {
    "tier": ("tier",),
    "region": ("region",),
    "partnership_scope": ("region", "scope"),
    "partnership_type": ("type", "partnership type"),
    "asset_class": ("asset class",),
    "assigned_owner": ("owner", "blk contact", "assigned to"),
    "contact": ("primary contact", "contact", "contact on file"),
    "email": ("email(s)", "email", "email on file"),
    "expiration": ("expiration", "last cycle / expiration", "last cycle"),
    "renewal_notes": ("renewal notes",),
    "last_touchpoint": ("last touchpoint", "last touch", "last inbound"),
    "email_chain_notes": ("email chain notes (jul 2026 scan)",),
    "contact_verified_status": ("contact verified",),
    "next_step": ("next step",),
    "next_step_due": ("next step due",),
}

EXPLICIT_CORRECTIONS: dict[str, dict[str, str]] = {
    "jpmorgan": {
        "firm": "J.P. Morgan",
        "relationship_status": "Not Interested",
        "pipeline_stage": "Closed / No Active Workflow",
    },
    "sixth street": {
        "firm": "Sixth Street Partners",
        "relationship_status": "Global Partner",
        "pipeline_stage": "Closed / Partner",
    },
    "sentinel": {
        "firm": "Sentinel Partners",
        "relationship_status": "Expired / Former Partner",
        "pipeline_stage": "Re-engagement",
    },
    "harbourvest": {
        "firm": "HarbourVest",
        "relationship_status": "Expired / Former Partner",
        "pipeline_stage": "Re-engagement",
    },
    "advent": {
        "firm": "Advent International",
        "relationship_status": "Existing Partner",
    },
    "point72": {
        "firm": "Point72",
        "relationship_status": "Not Renewing",
        "pipeline_stage": "Re-engagement",
    },
}


def firm_key(value: Any) -> str:
    """Conservative firm identity plus a small reviewed alias set."""
    normalized = normalize_firm(str(value or ""))
    compact = re.sub(r"[^a-z0-9]", "", str(value or "").lower())
    if "jpmorgan" in compact or compact.startswith("jpmorganchase"):
        return "jpmorgan"
    if "sixthstreet" in compact:
        return "sixth street"
    if "sentinelpartners" in compact:
        return "sentinel"
    if "harbourvest" in compact:
        return "harbourvest"
    if "adventinternational" in compact:
        return "advent"
    if compact.startswith("point72"):
        return "point72"
    return normalized


def clean(value: Any) -> str:
    return " ".join(str(value or "").strip().split())


def header_index(headers: list[str], names: Iterable[str]) -> int | None:
    lowered = {header.lower(): index for index, header in enumerate(headers) if header}
    for name in names:
        if name in lowered:
            return lowered[name]
    return None


def find_header(rows: list[tuple[Any, ...]]) -> tuple[int, list[str]] | None:
    for index, row in enumerate(rows[:10]):
        headers = [clean(value) for value in row]
        if any(header.lower() in {"firm", "partnership"} for header in headers):
            return index, headers
    return None


@dataclass(frozen=True)
class SourceRow:
    sheet: str
    row_number: int
    firm: str
    values: dict[str, Any]

    @property
    def authority(self) -> int:
        return AUTHORITY[self.sheet]

    @property
    def source(self) -> str:
        return f"{self.sheet}!r{self.row_number}"


@dataclass(frozen=True)
class Candidate:
    value: str
    source: str
    sheet: str
    authority: int
    issue: str = ""


@dataclass(frozen=True)
class ProposedChange:
    firm: str
    field: str
    current_db: str
    proposed: str
    source: str
    authority: int
    category: str
    confidence_issue: str


def read_workbook(path: Path) -> tuple[list[SourceRow], dict[str, Any]]:
    """Read relevant sheets without saving or modifying the workbook."""
    import openpyxl

    before = path.stat()
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    source_rows: list[SourceRow] = []
    inventory: dict[str, Any] = {"workbook": path.name, "sheets": {}}
    try:
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            preview = [tuple(row) for row in sheet.iter_rows(min_row=1, max_row=10, values_only=True)]
            found = find_header(preview)
            if not found:
                inventory["sheets"][sheet_name] = {"read": False, "reason": "no firm header"}
                continue
            header_offset, headers = found
            firm_col = header_index(headers, ("firm", "partnership"))
            count = 0
            for row_number, raw in enumerate(
                sheet.iter_rows(min_row=header_offset + 2, values_only=True),
                start=header_offset + 2,
            ):
                if firm_col is None or firm_col >= len(raw):
                    continue
                firm = clean(raw[firm_col])
                if not firm:
                    continue
                values = {
                    header: raw[index] if index < len(raw) else None
                    for index, header in enumerate(headers) if header
                }
                source_rows.append(SourceRow(sheet_name, row_number, firm, values))
                count += 1
            inventory["sheets"][sheet_name] = {
                "read": sheet_name in CORE_SHEETS,
                "rows": count,
                "headers": [header for header in headers if header],
            }
    finally:
        workbook.close()
    after = path.stat()
    inventory["workbook_unchanged"] = (
        before.st_size == after.st_size and before.st_mtime == after.st_mtime
    )
    return [row for row in source_rows if row.sheet in CORE_SHEETS], inventory


def read_database_csv(path: Path) -> list[dict[str, Any]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def read_database_json(path: Path) -> list[dict[str, Any]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(payload, dict):
        payload = payload.get("targets") or payload.get("data") or []
    if not isinstance(payload, list):
        raise ValueError("Database JSON must be a list or contain a targets list.")
    return [dict(row) for row in payload]


def date_candidate(value: Any) -> tuple[str, str]:
    if value in (None, ""):
        return "", ""
    if isinstance(value, datetime):
        return value.date().isoformat(), ""
    if isinstance(value, date):
        return value.isoformat(), ""
    if isinstance(value, (int, float)) and 1900 < int(value) < 2200:
        return str(int(value)), "year-only value; exact date not inferred"
    text = clean(value)
    if re.fullmatch(r"(19|20)\d{2}", text):
        return text, "year-only value; exact date not inferred"
    if re.search(r"\b(expired|lapsed|historical)\b", text, re.I):
        return text, "descriptive/ambiguous date retained as text"
    known = ("%b %d, %Y", "%B %d, %Y", "%b %d %Y", "%m/%d/%Y", "%Y-%m-%d")
    for fmt in known:
        try:
            return datetime.strptime(text, fmt).date().isoformat(), ""
        except ValueError:
            continue
    return text, "unparsed date retained as text"


def value_for(row: SourceRow, field: str) -> tuple[str, str]:
    values = {str(header).lower(): value for header, value in row.values.items()}
    if field == "relationship_status":
        if row.sheet in {"Pipeline & Leads", "Follow-Up Tracker"}:
            return "", ""
        raw = values.get("status") or ""
        return canonical_relationship(raw), ""
    if field == "pipeline_stage":
        raw = values.get("stage") or ""
        notes = " ".join(clean(values.get(name)) for name in (
            "notes", "renewal notes", "next step", "bucket"
        ))
        if row.sheet == "Follow-Up Tracker":
            bucket = clean(values.get("bucket"))
            if bucket:
                return "Follow-Up Due", "supporting evidence only"
            return "", ""
        if row.sheet == "Pipeline & Leads" or raw or "re-engag" in notes.lower():
            relationship = canonical_relationship(values.get("status"))
            return canonical_pipeline_stage(raw, relationship_status=relationship, notes=notes), ""
        return "", ""
    header_names = FIELD_HEADERS[field]
    value = next((values.get(header) for header in header_names if values.get(header) not in (None, "")), None)
    if field in {"expiration", "next_step_due"}:
        return date_candidate(value)
    return clean(value), ""


def explicit_candidates(key: str) -> dict[str, Candidate]:
    correction = EXPLICIT_CORRECTIONS.get(key, {})
    return {
        field: Candidate(value, "Explicit correction", "Explicit correction", 100)
        for field, value in correction.items() if field != "firm"
    }


def choose_candidates(rows: list[SourceRow], key: str) -> tuple[dict[str, Candidate], list[str]]:
    fields = (
        "relationship_status", "pipeline_stage", "tier", "region",
        "partnership_scope", "partnership_type", "asset_class", "assigned_owner",
        "contact", "email", "expiration", "renewal_notes", "last_touchpoint",
        "email_chain_notes", "contact_verified_status", "next_step", "next_step_due",
    )
    chosen = explicit_candidates(key)
    conflicts: list[str] = []
    for field in fields:
        if field in chosen:
            continue
        candidates: list[Candidate] = []
        for row in rows:
            value, issue = value_for(row, field)
            if value:
                candidates.append(Candidate(value, row.source, row.sheet, row.authority, issue))
        if not candidates:
            continue
        candidates.sort(key=lambda candidate: (-candidate.authority, candidate.source))
        highest = [
            candidate for candidate in candidates
            if candidate.authority == candidates[0].authority
        ]
        if len({candidate.value.casefold() for candidate in highest}) > 1:
            conflicts.append(
                f"{field}: equal-authority tie; " + "; ".join(
                    f"{candidate.value} ({candidate.source})" for candidate in highest[:5]
                )
            )
            continue
        chosen[field] = candidates[0]
        distinct = {candidate.value.casefold() for candidate in candidates}
        if len(distinct) > 1:
            conflicts.append(
                f"{field}: " + "; ".join(
                    f"{candidate.value} ({candidate.source})" for candidate in candidates[:5]
                )
            )
    return chosen, conflicts


def database_record(row: Mapping[str, Any]) -> dict[str, str]:
    relationship = str(row.get("relationship_status_effective") or "").strip() or (
        str(row.get("relationship_status_override") or "").strip()
        or str(row.get("relationship_status_auto") or "").strip()
        or canonical_relationship(row.get("relationship_status"), row.get("contact_status"))
    )
    stage = str(row.get("pipeline_stage_effective") or "").strip() or (
        str(row.get("pipeline_stage_override") or "").strip()
        or str(row.get("pipeline_stage_auto") or "").strip()
        or canonical_pipeline_stage(
            row.get("crm_status") or row.get("status"),
            relationship_status=relationship,
            notes=row.get("notes"),
        )
    )
    return {
        "firm": clean(row.get("firm")),
        "relationship_status": relationship,
        "pipeline_stage": stage,
        "tier": clean(row.get("sponsorship_tier") or row.get("relationship_tier") or row.get("tier_target")),
        "region": clean(row.get("region")),
        "partnership_scope": clean(row.get("partnership_scope") or row.get("region")),
        "partnership_type": clean(row.get("partnership_type") or row.get("relationship")),
        "asset_class": clean(row.get("firm_type")),
        "assigned_owner": clean(row.get("assigned_owner") or row.get("owner")),
        "contact": clean(row.get("relationship_contact_name")),
        "email": clean(row.get("relationship_contact_email")),
        "expiration": clean(row.get("relationship_expiration")),
        "renewal_notes": clean(row.get("renewal_notes") or row.get("relationship_decline_reason")),
        "last_touchpoint": clean(row.get("last_touchpoint")),
        "email_chain_notes": clean(row.get("email_chain_notes")),
        "contact_verified_status": clean(row.get("contact_verified_status")),
        "next_step": clean(row.get("next_step")),
        "next_step_due": clean(row.get("next_step_due")),
        "domain": clean(row.get("domain")),
    }


def ambiguous_db_match(key: str, database: dict[str, dict[str, str]]) -> tuple[str, float] | None:
    scored = sorted(
        ((candidate, SequenceMatcher(None, key, candidate).ratio()) for candidate in database),
        key=lambda item: item[1], reverse=True,
    )
    if scored and scored[0][1] >= 0.84:
        return scored[0]
    return None


def reconcile(
    source_rows: list[SourceRow], database_rows: list[dict[str, Any]]
) -> dict[str, Any]:
    database = {
        firm_key(row.get("firm")): database_record(row)
        for row in database_rows if firm_key(row.get("firm"))
    }
    database_ids = {
        clean(row.get("relationship_record_id")).casefold(): firm_key(row.get("firm"))
        for row in database_rows if clean(row.get("relationship_record_id"))
    }
    database_domains = {
        clean(row.get("domain")).casefold().removeprefix("www."): firm_key(row.get("firm"))
        for row in database_rows if clean(row.get("domain"))
    }
    workbook_by_key: dict[str, list[SourceRow]] = defaultdict(list)
    for row in source_rows:
        values = {str(header).casefold(): value for header, value in row.values.items()}
        record_id = clean(values.get("id")).casefold()
        domain = clean(values.get("domain") or values.get("website")).casefold()
        domain = re.sub(r"^https?://", "", domain).split("/", 1)[0].removeprefix("www.")
        key = database_ids.get(record_id) or database_domains.get(domain) or firm_key(row.firm)
        if key:
            workbook_by_key[key].append(row)

    # Explicit corrections participate even when a workbook row is absent.
    for key in EXPLICIT_CORRECTIONS:
        workbook_by_key.setdefault(key, [])

    changes: list[ProposedChange] = []
    firm_results: list[dict[str, Any]] = []
    exact_matches = 0
    new_firms = 0
    ambiguous_matches = 0
    explicit_cases = 0
    manual_review: list[dict[str, Any]] = []

    duplicates = [
        {
            "firm_key": key,
            "sheet": sheet,
            "rows": [row.row_number for row in rows if row.sheet == sheet],
        }
        for key, rows in workbook_by_key.items()
        for sheet, count in Counter(row.sheet for row in rows).items()
        if count > 1
    ]

    for key, rows in sorted(workbook_by_key.items()):
        chosen, conflicts = choose_candidates(rows, key)
        db = database.get(key)
        issue_parts: list[str] = []
        if conflicts:
            issue_parts.append("conflicting sources: " + " | ".join(conflicts))
        if db is None:
            near = ambiguous_db_match(key, database)
            if near:
                ambiguous_matches += 1
                issue_parts.append(f"ambiguous DB match: {database[near[0]]['firm']} ({near[1]:.0%})")
            else:
                new_firms += 1
        if key in EXPLICIT_CORRECTIONS:
            explicit_cases += 1

        firm_name = (
            EXPLICIT_CORRECTIONS.get(key, {}).get("firm")
            or (rows[0].firm if rows else key)
        )
        proposed_values = {field: candidate.value for field, candidate in chosen.items()}
        changed = 0
        for field, candidate in chosen.items():
            current = db.get(field, "") if db else ""
            if clean(current).casefold() == clean(candidate.value).casefold():
                continue
            category = (
                "explicit user override" if candidate.sheet == "Explicit correction"
                else "new firm" if db is None
                else "conflicting source" if conflicts
                else "safe update"
            )
            confidence = (
                "explicit instruction; highest authority"
                if candidate.sheet == "Explicit correction"
                else candidate.issue or (
                    "requires manual review" if issue_parts else "high"
                )
            )
            changes.append(ProposedChange(
                firm=firm_name,
                field=field,
                current_db=current,
                proposed=candidate.value,
                source=candidate.source,
                authority=candidate.authority,
                category=category,
                confidence_issue=confidence,
            ))
            changed += 1
        if db is not None and changed == 0:
            exact_matches += 1
        result = {
            "firm_key": key,
            "firm": firm_name,
            "database_match": db.get("firm") if db else None,
            "proposed": proposed_values,
            "conflicts": conflicts,
            "issues": issue_parts,
            "categories": [
                *(["new firm", "missing DB record"] if db is None else []),
                *(["conflicting source"] if conflicts else []),
                *(["ambiguous match"] if any("ambiguous DB match" in issue for issue in issue_parts) else []),
                *(["explicit user override"] if key in EXPLICIT_CORRECTIONS else []),
            ],
            "source_rows": [row.source for row in rows],
        }
        firm_results.append(result)
        if issue_parts or conflicts or db is None:
            manual_review.append(result)

    workbook_keys = set(workbook_by_key)
    missing_workbook = [
        row for key, row in sorted(database.items()) if key not in workbook_keys
    ]
    proposals_by_key = {
        result["firm_key"]: result["proposed"] for result in firm_results
    }
    missing_workbook_values = [
        {
            "firm": database[key]["firm"],
            "field": field,
            "retained_database_value": database[key][field],
            "category": "missing workbook value",
        }
        for key in sorted(workbook_keys & set(database))
        for field in (
            "relationship_status", "pipeline_stage", "tier", "region",
            "partnership_scope", "partnership_type", "asset_class", "assigned_owner",
            "contact", "email", "expiration", "renewal_notes", "last_touchpoint",
            "email_chain_notes", "contact_verified_status", "next_step", "next_step_due",
        )
        if database[key].get(field) and field not in proposals_by_key[key]
    ]
    summary = {
        "workbook_records_inspected": len(source_rows),
        "unique_workbook_firms": len(workbook_by_key),
        "database_records_inspected": len(database_rows),
        "exact_matches": exact_matches,
        "proposed_updates": sum(1 for change in changes if change.category != "new firm"),
        "safe_updates": sum(change.category == "safe update" for change in changes),
        "new_firms": new_firms,
        "missing_database_records": new_firms,
        "conflicts": sum(bool(result["conflicts"]) for result in firm_results),
        "ambiguous_matches": ambiguous_matches,
        "duplicates": len(duplicates),
        "explicit_override_cases": explicit_cases,
        "records_requiring_manual_review": len(manual_review),
        "missing_workbook_values": len(missing_workbook_values),
        "database_records_missing_from_workbook": len(missing_workbook),
        "database_mutations": 0,
    }
    return {
        "summary": summary,
        "changes": [asdict(change) for change in changes],
        "firms": firm_results,
        "duplicates": duplicates,
        "manual_review": manual_review,
        "missing_workbook_values": missing_workbook_values,
        "database_records_missing_from_workbook": missing_workbook,
    }


def write_reports(result: dict[str, Any], inventory: dict[str, Any], output_dir: Path) -> dict[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    json_path = output_dir / "crm_reconciliation_dry_run.json"
    csv_path = output_dir / "crm_reconciliation_changes.csv"
    markdown_path = output_dir / "crm_reconciliation_summary.md"
    payload = {"dry_run": True, "inventory": inventory, **result}
    json_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")

    fields = [
        "firm", "field", "current_db", "proposed", "source", "authority",
        "category", "confidence_issue",
    ]
    with csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(result["changes"])

    summary = result["summary"]
    priority = sorted(
        result["changes"],
        key=lambda row: (-int(row["authority"]), row["firm"], row["field"]),
    )[:30]
    lines = [
        "# CRM reconciliation dry run",
        "",
        "> Review only. No workbook or database records were changed.",
        "",
        "## Summary",
        "",
        *[f"- {key.replace('_', ' ').title()}: {value}" for key, value in summary.items()],
        "",
        "## Highest-priority proposed changes",
        "",
        "| Firm | Field | Current DB | Proposed | Source | Authority | Confidence / issue |",
        "| --- | --- | --- | --- | --- | ---: | --- |",
    ]
    for row in priority:
        cells = [
            str(row[name]).replace("|", "\\|")
            for name in ("firm", "field", "current_db", "proposed", "source")
        ]
        issue = str(row["confidence_issue"]).replace("|", "\\|")
        lines.append(f"| {' | '.join(cells)} | {row['authority']} | {issue} |")
    markdown_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return {"json": json_path, "csv": csv_path, "markdown": markdown_path}


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--database-json", type=Path)
    parser.add_argument("--database-csv", type=Path, default=DEFAULT_DATABASE_CSV)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    args = parser.parse_args(argv)

    if not args.workbook.exists():
        print(f"Workbook not found: {args.workbook}", file=sys.stderr)
        return 2
    database_source = args.database_json or args.database_csv
    if not database_source.exists():
        print(f"Database snapshot not found: {database_source}", file=sys.stderr)
        return 2

    source_rows, inventory = read_workbook(args.workbook)
    database_rows = (
        read_database_json(args.database_json)
        if args.database_json else read_database_csv(args.database_csv)
    )
    inventory["database_source"] = str(database_source)
    inventory["database_source_kind"] = (
        "database_json_export" if args.database_json else "reviewed_local_seed_snapshot"
    )
    result = reconcile(source_rows, database_rows)
    reports = write_reports(result, inventory, args.output_dir)
    print(json.dumps(result["summary"], indent=2))
    print("DRY RUN: zero database changes; workbook unchanged = " + str(inventory["workbook_unchanged"]))
    for label, path in reports.items():
        print(f"{label}: {path}")
    return 0 if inventory["workbook_unchanged"] else 3


if __name__ == "__main__":
    raise SystemExit(main())
